import pandas as pd
import os
import openpyxl
import zipfile
import numpy as np

def load_excel(file_path):
    try:
        return pd.read_excel(file_path, engine='openpyxl', sheet_name=None)
    except (FileNotFoundError, zipfile.BadZipFile, ValueError) as e:
        print(f"Ошибка при чтении файла {file_path}: {e}")
        return None

# Обрабокта файла с трафаретами
directory_path = '/Users/gen/PycharmProjects/Files/pattern/'
pattern = pd.DataFrame(columns=['SKU', 'Расход, ₽, с НДС'])
files_in_directory = os.listdir(directory_path)

for filename in files_in_directory:
    file_path = os.path.join(directory_path, filename)

    if os.path.isfile(file_path) and filename.endswith('.xlsx'):
        try:
            # Открываем файл Excel
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            # Выбираем активный лист (первый лист)
            sheet = workbook.active

            # Собираем строки для удаления
            rows_to_delete = []

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and ("Кампания по продвижению товаров" in str(cell.value) or cell.value == "Всего"):
                        rows_to_delete.append(cell.row)
                        break

            # Удаляем строки в обратном порядке
            for row_num in sorted(rows_to_delete, reverse=True):
                sheet.delete_rows(row_num)

            # Сохраняем изменения и закрываем файл Excel
            workbook.save(file_path)
            workbook.close()

            # Загружаем данные из измененного файла и добавляем их в переменную pattern
            df = load_excel(file_path)
            if df is not None:
                for sheet_name, sheet_data in df.items():
                    if not pattern.empty:
                        pattern = pd.concat([pattern, sheet_data], ignore_index=True)
                    else:
                        pattern = sheet_data

        except Exception as e:
            continue

# *** Обработка Брендовая полка ***
directory_path = '/Users/gen/PycharmProjects/Files/shelf/'

# Создаем пустой DataFrame, в который будем объединять данные
df_shelf = pd.DataFrame()

# Перебор всех файлов в папке
for filename in os.listdir(directory_path):
    full_path = os.path.join(directory_path, filename)

    # Проверка, что это файл xlsx
    if os.path.isfile(full_path) and filename.endswith('.xlsx'):
        try:
            # Чтение данных из файла Excel в DataFrame, начиная с второй строки
            df = pd.read_excel(full_path, engine='openpyxl', header=1)

            # Проверяем, существует ли столбец "Расход, ₽, с НДС" в файле
            if 'Расход, ₽, с НДС' in df.columns:
                # Добавляем столбец "Расход, ₽, с НДС" в объединенные данные
                df_shelf = pd.concat([df_shelf, df['Расход, ₽, с НДС']], ignore_index=True)
                print(f'Данные из столбца "Расход, ₽, с НДС" в файле {filename} добавлены в объединенные данные.')
            else:
                print(f'В файле {filename} отсутствует столбец "Расход, ₽, с НДС".')

        except Exception as e:
            print(f"Ошибка при обработке файла {full_path}: {e}")

# Сохраняем объединенные данные в новый файл Excel с указанием названия первой строки
combined_file_path = '/Users/gen/PycharmProjects/Files/shelf/combined_data.xlsx'
df_shelf.to_excel(combined_file_path, index=False, engine='openpyxl', header=['Расход, ₽, с НДС'])
print(f'Объединенные данные сохранены в файле {combined_file_path}.')

# *** Загрузка и обработка файлов Продвижение в поиске ***
directory_path = '/Users/gen/PycharmProjects/Files/search/'  # Указываем путь к директории с файлами

search = pd.DataFrame(columns=['SKU', 'Расход, ₽, с НДС'])  # Создаем пустой DataFrame для хранения данных

files_in_directory = os.listdir(directory_path)  # Получаем список файлов в директории

for filename in files_in_directory:  # Проходим по каждому файлу
    file_path = os.path.join(directory_path, filename)

    if os.path.isfile(file_path) and filename.endswith('.xlsx'):  # Проверяем, что файл существует и имеет расширение .xlsx
        workbook_search = openpyxl.load_workbook(file_path)  # Открываем файл Excel

        sheet_search = workbook_search.active  # Выбираем активный лист (первый лист)

        for row_search in sheet_search.iter_rows():  # Ищем строки, содержащие слова "Кампания по продвижению товаров" и удаляем их
            for cell_search in row_search:
                if cell_search.value and "Кампания по продвижению товаров" in str(cell_search.value):
                    sheet_search.delete_rows(cell_search.row)
                    break  # Выходим из цикла, чтобы не удалять несколько строк с одним и тем же значением

        for row_search in sheet_search.iter_rows():  # Ищем строки, содержащие слово "Всего" и удаляем их
            for cell_search in row_search:
                if cell_search.value == "Всего":
                    sheet_search.delete_rows(cell_search.row)
                    break  # Выходим из цикла, чтобы не удалять несколько строк с одним и тем же значением

        workbook_search.save(file_path)  # Закрываем файл Excel и сохраняем изменения
        workbook_search.close()

        df_search = load_excel(file_path)  # Загружаем данные из измененного файла и добавляем их в переменную search
        if df_search is not None:
            for sheet_name_search, sheet_data_search in df_search.items():
                if not search.empty:
                    search = pd.concat([search, sheet_data_search], ignore_index=True)
                else:
                    search = sheet_data_search

# Теперь переменная search содержит объединенные данные из всех файлов Excel без первой строки
print(search)













# Загружаем файлы по прямому пути
direct_paths = ['/Users/gen/PycharmProjects/Files/net3.xlsx',]

# Если вы хотите что-то сделать с первым листом из конкретного файла
excel_data = load_excel('/Users/gen/PycharmProjects/Files/net3.xlsx')
if excel_data:
    first_sheet_name = list(excel_data.keys())[0]
    df1 = excel_data[first_sheet_name]
    # Теперь вы можете работать с df1, если это необходимо

# Переменная со всеми необходимыми столбцами
required_columns = [
    "SKU", "Артикул", "Название товара или услуги", "Количество отменных товаров в шт.", "Выкуплено товаров в шт.",
    "Выкуплено товаров", "Комиссия за продажу", "Последняя миля", "Обработка отправления «Drop-off»", "Логистика",
    "Получено возвратов", "Возврат комиссии", "Обработка отмененных и невостребованных товаров",
    "Обработка частичного невыкупа", "Обратная логистика", "Обработка возврата", "Услуга по изменению условий отгрузки",
    "Premium-подписка", "Транспортно-экспедиционные услуги Кроссдокинг", "Услуга Гибкий график выплат",
    "Приобретение отзывов на платформе", "Услуга размещения товаров", "Утилизация со стока", "Оплата эквайринга", "Расход Продвижение в поиске", "Расход Трафареты", "Расход Брендовая полка"
]

df2 = pd.DataFrame(columns=required_columns)








# Обработка импортированных файлов
for imported_df in [df1]:
    if imported_df is not None:  # Добавьте проверку на None, чтобы избежать ошибки
        missing_columns = [col for col in required_columns if col not in imported_df.columns]

        if missing_columns:
            # Если в импортированном файле отсутствуют столбцы, заполняем их нулями
            for col in missing_columns:
                df2[col] = 0
        else:
            # В противном случае, копируем данные из импортированного файла в df2
            df2 = pd.concat([df2, imported_df[required_columns]], ignore_index=True)

#Группировка столбцов 'SKU', 'Артикул', 'Название товара или услуги' из переменной df1 и перенос их в переменную df2
grouped_df1 = df1.groupby('SKU').first().reset_index()[['SKU', 'Артикул', 'Название товара или услуги']]
df2[['SKU', 'Артикул', 'Название товара или услуги']] = grouped_df1

# Обработка 'SKU', 'Артикул', 'Название товара или услуги'
if 'SKU' in df1.columns and 'Артикул' in df1.columns and 'Название товара или услуги' in df1.columns:
    grouped_df1 = df1.groupby('SKU').first().reset_index()[['SKU', 'Артикул', 'Название товара или услуги']]
    df2[['SKU', 'Артикул', 'Название товара или услуги']] = grouped_df1
else:
    df2['SKU'] = 0
    df2['Артикул'] = 0
    df2['Название товара или услуги'] = 0





# *** Заказы ***

# Количество отменных товаров в шт.
if 'Тип начисления' in df1.columns and 'Количество' in df1.columns:
    filtered_df1 = df1[df1['Тип начисления'] == 'Получение возврата, отмены, невыкупа от покупателя']
    grouped = filtered_df1.groupby('SKU')['Количество'].sum().reset_index()
    grouped.set_index('SKU', inplace=True)
    df2['Количество отменных товаров в шт.'] = df2['SKU'].map(grouped['Количество']).fillna(0)
else:
    df2['Количество отменных товаров в шт.'] = 0

# Выкуплено товаров в шт.
if 'Тип начисления' in df1.columns and 'За продажу или возврат до вычета комиссий и услуг' in df1.columns:
    filtered_df1 = df1[df1['Тип начисления'] == 'Доставка покупателю']
    grouped = filtered_df1.groupby('SKU')['Количество'].sum().reset_index()
    grouped.set_index('SKU', inplace=True)
    df2['Выкуплено товаров в шт.'] = df2['SKU'].map(grouped['Количество']).fillna(0)
else:
    df2['Выкуплено товаров в шт.'] = 0

# Выкуплено товаров
if 'Тип начисления' in df1.columns and 'За продажу или возврат до вычета комиссий и услуг' in df1.columns:
    filtered_df1 = df1[df1['Тип начисления'] == 'Доставка покупателю']
    grouped = filtered_df1.groupby('SKU')['За продажу или возврат до вычета комиссий и услуг'].sum().reset_index()
    grouped.set_index('SKU', inplace=True)
    df2['Выкуплено товаров'] = df2['SKU'].map(grouped['За продажу или возврат до вычета комиссий и услуг']).fillna(0)
else:
    df2['Выкуплено товаров'] = 0





# Комиссия за продажу
if 'Тип начисления' in df1.columns and 'Комиссия за продажу' in df1.columns:
    filtered_df1 = df1[df1['Тип начисления'] == 'Доставка покупателю']
    grouped = filtered_df1.groupby('SKU')['Комиссия за продажу'].sum().reset_index()
    grouped.set_index('SKU', inplace=True)
    df2['Комиссия за продажу'] = df2['SKU'].map(grouped['Комиссия за продажу']).fillna(0)
else:
    df2['Комиссия за продажу'] = 0

# Последняя миля
if 'Тип начисления' in df1.columns and 'Последняя миля (разбивается по товарам пропорционально доле цены товара в сумме отправления)' in df1.columns:
    filtered_df1 = df1[df1['Тип начисления'] == 'Доставка покупателю']
    grouped = filtered_df1.groupby('SKU')['Последняя миля (разбивается по товарам пропорционально доле цены товара в сумме отправления)'].sum().reset_index()
    grouped.set_index('SKU', inplace=True)
    df2['Последняя миля'] = df2['SKU'].map(grouped['Последняя миля (разбивается по товарам пропорционально доле цены товара в сумме отправления)']).fillna(0)
else:
    df2['Последняя миля'] = 0

# Обработка отправления «Drop-off»
if 'Тип начисления' in df1.columns and 'Обработка отправления (Drop-off/Pick-up) (разбивается по товарам пропорционально количеству в отправлении)' in df1.columns:
    filtered_df1 = df1[df1['Тип начисления'] == 'Доставка покупателю']
    grouped = filtered_df1.groupby('SKU')['Обработка отправления (Drop-off/Pick-up) (разбивается по товарам пропорционально количеству в отправлении)'].sum().reset_index()
    grouped.set_index('SKU', inplace=True)
    df2['Обработка отправления «Drop-off»'] = df2['SKU'].map(grouped['Обработка отправления (Drop-off/Pick-up) (разбивается по товарам пропорционально количеству в отправлении)']).fillna(0)
else:
    df2['Обработка отправления «Drop-off»'] = 0

# Логистика
if 'Тип начисления' in df1.columns and 'Логистика' in df1.columns:
    # Добавляем фильтрацию по двум условиям
    filtered_df1 = df1[df1['Тип начисления'].isin(['Доставка покупателю', 'Доставка и обработка возврата, отмены, невыкупа'])]
    grouped = filtered_df1.groupby('SKU')['Логистика'].sum().reset_index()
    grouped.set_index('SKU', inplace=True)
    df2['Логистика'] = df2['SKU'].map(grouped['Логистика']).fillna(0)
else:
    df2['Логистика'] = 0

# *** Возвраты и отмены ***
# Получено возвратов
if 'Тип начисления' in df1.columns and 'За продажу или возврат до вычета комиссий и услуг' in df1.columns:
    filtered_df1 = df1[df1['Тип начисления'] == 'Получение возврата, отмены, невыкупа от покупателя']
    grouped = filtered_df1.groupby('SKU')['За продажу или возврат до вычета комиссий и услуг'].sum().reset_index()
    grouped.set_index('SKU', inplace=True)
    df2['Получено возвратов'] = df2['SKU'].map(grouped['За продажу или возврат до вычета комиссий и услуг']).fillna(0)
else:
    df2['Получено возвратов'] = 0

# Возврат комиссии
if 'Тип начисления' in df1.columns and 'Комиссия за продажу' in df1.columns:
    filtered_df1 = df1[df1['Тип начисления'] == 'Получение возврата, отмены, невыкупа от покупателя']
    grouped = filtered_df1.groupby('SKU')['Комиссия за продажу'].sum().reset_index()
    grouped.set_index('SKU', inplace=True)
    df2['Возврат комиссии'] = df2['SKU'].map(grouped['Комиссия за продажу']).fillna(0)
else:
    df2['Возврат комиссии'] = 0

# Обработка отмененных и невостребованных товаров
if 'Тип начисления' in df1.columns and 'Обработка отмененного или невостребованного товара (разбивается по товарам в отправлении в одинаковой пропорции)' in df1.columns:
    filtered_df1 = df1[df1['Тип начисления'] == 'Доставка и обработка возврата, отмены, невыкупа']
    grouped = filtered_df1.groupby('SKU')['Обработка отмененного или невостребованного товара (разбивается по товарам в отправлении в одинаковой пропорции)'].sum().reset_index()
    grouped.set_index('SKU', inplace=True)
    df2['Обработка отмененных и невостребованных товаров'] = df2['SKU'].map(grouped['Обработка отмененного или невостребованного товара (разбивается по товарам в отправлении в одинаковой пропорции)']).fillna(0)
else:
    df2['Обработка отмененных и невостребованных товаров'] = 0

# Обработка частичного невыкупа
if 'Тип начисления' in df1.columns and 'Обработка невыкупленного товара' in df1.columns:
    filtered_df1 = df1[df1['Тип начисления'] == 'Доставка и обработка возврата, отмены, невыкупа']
    grouped = filtered_df1.groupby('SKU')['Обработка невыкупленного товара'].sum().reset_index()
    grouped.set_index('SKU', inplace=True)
    df2['Обработка частичного невыкупа'] = df2['SKU'].map(grouped['Обработка невыкупленного товара']).fillna(0)
else:
    df2['Обработка частичного невыкупа'] = 0

# Обратная логистика
if 'Тип начисления' in df1.columns and 'За продажу или возврат до вычета комиссий и услуг' in df1.columns:
    filtered_df1 = df1[df1['Тип начисления'] == 'Доставка и обработка возврата, отмены, невыкупа']
    grouped = filtered_df1.groupby('SKU')['Обратная логистика'].sum().reset_index()
    grouped.set_index('SKU', inplace=True)
    df2['Обратная логистика'] = df2['SKU'].map(grouped['Обратная логистика']).fillna(0)
else:
    df2['Обратная логистика'] = 0

# Обработка возвратов
if 'Тип начисления' in df1.columns and 'Обработка возврата' in df1.columns:
    filtered_df1 = df1[df1['Тип начисления'] == 'Доставка и обработка возврата, отмены, невыкупа']
    grouped = filtered_df1.groupby('SKU')['Обработка возврата'].sum().reset_index()
    grouped.set_index('SKU', inplace=True)
    df2['Обработка возврата'] = df2['SKU'].map(grouped['Обработка возврата']).fillna(0)
else:
    df2['Обработка возврата'] = 0

# *** Услуги ***
# Заполните весь столбец "Premium-подписка" в df2 нулями
df2['Premium-подписка'] = 0
# Выберите значение из столбца "Итого" на основе условия "Тип начисления" == "Premium-подписка" в df1
value = df1.loc[df1['Тип начисления'] == 'Premium-подписка', 'Итого'].iloc[0]
# Установите это значение в первую ячейку столбца "Premium-подписка" в df2
df2.at[0, 'Premium-подписка'] = value

# Транспортно-экспедиционные услуги Кроссдокинг
# Заполните весь столбец "Транспортно-экспедиционные услуги Кроссдокинг" в df2 нулями
df2['Транспортно-экспедиционные услуги Кроссдокинг'] = 0
# Выберите значение из столбца "Итого" на основе условия "Тип начисления" == "Доставка товаров на склад Ozon (кросс-докинг)" в df1
value = df1.loc[df1['Тип начисления'] == 'Доставка товаров на склад Ozon (кросс-докинг)', 'Итого'].iloc[0]
# Установите это значение в первую ячейку столбца "Транспортно-экспедиционные услуги Кроссдокинг" в df2
df2.at[0, 'Транспортно-экспедиционные услуги Кроссдокинг'] = value

# Услуга Гибкий график выплат
# Заполните весь столбец "Услуга Гибкий график выплат" в df2 нулями
df2['Услуга Гибкий график выплат'] = 0
# Проверьте наличие строки с условием "Тип начисления" == "Начисление за гибкий график выплат" в df1
filtered_values = df1.loc[df1['Тип начисления'] == 'Начисление за гибкий график выплат', 'Итого']
if not filtered_values.empty:
    # Выберите значение из столбца "Итого" на основе данного условия
    value = filtered_values.iloc[0]
    df2['Услуга Гибкий график выплат'] = df2['Услуга Гибкий график выплат'].astype(float)
    df2.at[0, 'Услуга Гибкий график выплат'] = value

# Кросс-докинг
# Выборка строк из df1 и суммирование значений столбца "Итого"
total_sum = df1[df1['Тип начисления'] == 'Доставка товаров на склад Ozon (кросс-докинг)']['Итого'].sum()
# Присваивание суммы первой строке столбца "Услуга размещения товаров" в df2
df2.at[0, 'Транспортно-экспедиционные услуги Кроссдокинг'] = total_sum
# Заполнение всех остальных строк этого столбца в df2 нулями
df2.loc[1:, 'Транспортно-экспедиционные услуги Кроссдокинг'] = 0

# Приобретение отзывов на платформе
total_sum = df1[df1['Тип начисления'] == 'Приобретение отзывов на платформе']['Итого'].sum()
# Присваивание суммы первой строке столбца "Услуга размещения товаров" в df2
df2.at[0, 'Приобретение отзывов на платформе'] = total_sum
# Заполнение всех остальных строк этого столбца в df2 нулями
df2.loc[1:, 'Приобретение отзывов на платформе'] = 0

# Услуга размещения товаров
# Выборка строк из df1 и суммирование значений столбца "Итого"
total_sum = df1[df1['Тип начисления'] == 'Услуга размещения товаров на складе']['Итого'].sum()
# Присваивание суммы первой строке столбца "Услуга размещения товаров" в df2
df2.at[0, 'Услуга размещения товаров'] = total_sum
# Заполнение всех остальных строк этого столбца в df2 нулями
df2.loc[1:, 'Услуга размещения товаров'] = 0

# Заполните весь столбец "Утилизация со стока" в df2 нулями
df2['Утилизация со стока'] = 0
# Выберите значение из столбца "Итого" на основе условия "Тип начисления" == "Утилизация со стока" в df1
value = df1.loc[df1['Тип начисления'] == 'Утилизация', 'Итого'].iloc[0]
# Установите это значение в первую ячейку столбца "Утилизация со стока" в df2
df2.at[0, 'Утилизация со стока'] = value

# *** Компенсации и прочие начисления ***
# Выборка строк из df1 и суммирование значений столбца "Итого"
if 'Тип начисления' in df1.columns and 'Итого' in df1.columns:
    filtered_df1 = df1[df1['Тип начисления'] == 'Оплата эквайринга']
    grouped = filtered_df1.groupby('SKU')['Итого'].sum().reset_index()
    grouped.set_index('SKU', inplace=True)
    df2['Оплата эквайринга'] = df2['SKU'].map(grouped['Итого']).fillna(0)
else:
    df2['Оплата эквайринга'] = 0

# Услуга по изменению условий отгрузки
# Процесс 1
process_1 = df1[(df1['Тип начисления'] == 'Услуга по изменению условий отгрузки')][['Тип начисления', 'Номер отправления или идентификатор услуги', 'Итого']]
# Процесс 2
process_2 = df1[['Номер отправления или идентификатор услуги', 'SKU']]
# Процесс 3
# Объединяем процесс 1 и процесс 2 по столбцу 'Номер отправления или идентификатор услуги'
result = process_1.merge(process_2, on='Номер отправления или идентификатор услуги', how='left')
# Процесс 4
# Удаляем строки, где 'SKU' равно NaN
result = result.dropna(subset=['SKU'])



# Шаг 1: Выбрать столбцы 'Тип начисления', 'SKU' и 'Итого' из переменной result
result_selection = result[['Тип начисления', 'SKU', 'Итого']]

# Шаг 2: Выбрать столбец 'SKU' из переменной df2
df2_sku = df2['SKU']

# Шаг 3: Создать новый столбец 'Услуга по изменению условий отгрузки' в df2 и заполнить нулями
df2['Услуга по изменению условий отгрузки'] = 0

# Шаг 4: Сопоставить значения и перенести 'Итого' из result_selection в df2
for index, row in result_selection.iterrows():
    sku = row['SKU']
    total = row['Итого']
    # Проверить, что SKU совпадает и не NaN
    if not pd.isna(sku) and sku in df2_sku.values:
        # Найти индекс SKU в df2
        df2_index = df2[df2['SKU'] == sku].index[0]
        # Присвоить значение 'Итого' переменной 'Услуга по изменению условий отгрузки' в df2,
        # явно приведя total к целочисленному типу данных
        df2.at[df2_index, 'Услуга по изменению условий отгрузки'] = int(total)

# Шаг 5: Заполнить остальные ячейки 'Услуга по изменению условий отгрузки' в df2 нулями
df2['Услуга по изменению условий отгрузки'].fillna(0, inplace=True)

# *** Запись значения Pattern в общий файл ***
pattern.rename(columns={'sku': 'SKU'}, inplace=True)
# Убираем пустые значения в столбце SKU в переменной pattern
pattern = pattern.dropna(subset=['SKU'])
# Фильтруем SKU из pattern, которых нет в df2, и добавляем их в df2
new_skus = pattern[~pattern['SKU'].isin(df2['SKU'])]['SKU']
df2 = pd.concat([df2, pd.DataFrame({'SKU': new_skus})], ignore_index=True)
# Группируем по столбцу 'SKU' и суммируем значения 'Расход, ₽, с НДС'
pattern = pattern.groupby('SKU')['Расход, ₽, с НДС'].sum().reset_index()
pattern['Расход, ₽, с НДС'] = pattern['Расход, ₽, с НДС'] * -1  # Делаем значения отрицательными
# Сопоставляем значения 'SKU' из переменной pattern с df2 и записываем соответствующие значения 'Расход, ₽, с НДС'
df2['Расход Трафареты'] = df2['SKU'].map(pattern.set_index('SKU')['Расход, ₽, с НДС'])


# Запись значения Search в общий файл
search.rename(columns={'Ozon ID': 'SKU'}, inplace=True)  # Переименовываем столбец 'Ozon ID' в 'SKU'
search.rename(columns={'Расход Продвижение в поиске': 'Расход, ₽'}, inplace=True)  # Переименовываем столбец 'Расход Продвижение в поиске'
search = search.dropna(subset=['SKU'])  # Убираем строки с пустыми значениями в столбце 'SKU'
new_skus = search[~search['SKU'].isin(df2['SKU'])]['SKU']  # Фильтруем 'SKU' из search, которых нет в df2, и добавляем их в df2
df2 = pd.concat([df2, pd.DataFrame({'SKU': new_skus})], ignore_index=True)
search = search.groupby('SKU')['Расход, ₽'].sum().reset_index()  # Группируем по столбцу 'SKU' и суммируем значения 'Расход, ₽'
search['Расход, ₽'] = search['Расход, ₽'] * -1  # Делаем значения отрицательными
df2['Расход Продвижение в поиске'] = df2['SKU'].map(search.set_index('SKU')['Расход, ₽'])











# Запись в общий файл данных Брендовой полки
# Суммируем все значения из df_shelf
total_sum = df_shelf.sum()
# Делаем значение отрицательным
total_sum = total_sum * -1
# Записываем полученную сумму в новый столбец df2
df2['Расход Брендовая полка'] = total_sum

# Если total_sum является Series, преобразуем его в скалярное значение
if isinstance(total_sum, pd.Series):
    total_sum = total_sum.item()

# Теперь присваиваем значение:
df2.loc[1, 'Расход Брендовая полка'] = total_sum








# *** Запись результатов ***
df1.to_excel('данные.xlsx', engine='openpyxl', index=False)
# Запись результатов в Excel-файл, включая заголовки столбцов
df2.to_excel('данные2.xlsx', engine='openpyxl', index=False)
result.to_excel('данные3.xlsx', engine='openpyxl', index=False)
pattern.to_excel('pattern.xlsx', engine='openpyxl', index=False)
search.to_excel('search.xlsx', engine='openpyxl', index=False)







